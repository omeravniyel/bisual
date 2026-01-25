from fastapi import APIRouter, UploadFile, File, Depends, HTTPException, status, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.database import get_db
from app import models, schemas
import openpyxl
from io import BytesIO
import json

router = APIRouter(
    prefix="/import",
    tags=["import"]
)

TEMPLATE_HEADERS = [
    "Soru Metni", 
    "Süre (sn)", 
    "Puan (1000/2000/0)", 
    "A Seçeneği", 
    "B Seçeneği", 
    "C Seçeneği", 
    "D Seçeneği", 
    "Doğru Cevap (A/B/C/D)"
]

@router.get("/template")
def get_template():
    """Generates and returns a sample Excel template for quiz import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Soru Şablonu"
    
    # Write Headers
    for col_num, header in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = openpyxl.styles.Font(bold=True)
        # Adjust column width
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = 20

    # Write Sample Data
    sample_data = [
        ["Türkiye'nin başkenti neresidir?", 20, 1000, "İstanbul", "Ankara", "İzmir", "Bursa", "B"],
        ["Python dili hangi yıl çıkmıştır?", 30, 2000, "1989", "1991", "2000", "1995", "B"],
        ["Su kaç derecede kaynar?", 15, 1000, "100", "90", "80", "50", "A"]
    ]

    for row_idx, row_data in enumerate(sample_data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    return StreamingResponse(
        buffer, 
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=quiz_sablon.xlsx"}
    )

@router.post("/parse")
async def parse_quiz_excel(
    request: Request,
    file: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """Parses Excel and returns list of questions (no DB save)."""
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Sadece .xlsx dosyaları kabul edilir.")
        
    user_cookie = request.cookies.get("user_session")
    if not user_cookie: raise HTTPException(status_code=401, detail="Not authenticated")
    
    # Just verify existence
    user = db.query(models.User).filter(models.User.username == user_cookie).first()
    if not user: raise HTTPException(status_code=401, detail="User not found")

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(filename=BytesIO(contents), data_only=True)
        ws = wb.active
        
        header_row = [cell.value for cell in ws[1]]
        if not header_row or header_row[0] != TEMPLATE_HEADERS[0]:
            raise HTTPException(status_code=400, detail="Geçersiz şablon.")

        questions = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: continue
            
            q_text = str(row[0] or "").strip()
            time_limit = int(row[1] or 20)
            points = int(row[2] or 1000)
            
            opt_a = str(row[3] or "").strip()
            opt_b = str(row[4] or "").strip()
            opt_c = str(row[5] or "").strip()
            opt_d = str(row[6] or "").strip()
            
            correct_char = str(row[7] or "").strip().upper()
            mapping = {'A': 0, 'B': 1, 'C': 2, 'D': 3}
            correct_idx = mapping.get(correct_char, -1)
            
            if correct_idx == -1 and (opt_a or opt_b):
                 # For preview/parse, we can be slightly more lenient or just mark none, 
                 # but let's conform to the strict rule we just added or handle it on frontend.
                 # Let's return -1 and let frontend warn if needed, or just raise here too.
                 pass 

            questions.append({
                "text": q_text,
                "time_limit": time_limit,
                "points": points,
                "question_type": "multiple_choice",
                "image_url": None,
                "options": [
                    {"text": opt_a, "is_correct": correct_idx == 0},
                    {"text": opt_b, "is_correct": correct_idx == 1},
                    {"text": opt_c, "is_correct": correct_idx == 2},
                    {"text": opt_d, "is_correct": correct_idx == 3},
                ]
            })
            
        return questions

    except Exception as e:
        print(f"Parse Error: {e}")
        raise HTTPException(status_code=500, detail=f"Dosya okunamadı: {str(e)}")

@router.post("/upload")
async def import_quiz(
    request: Request,
    file: UploadFile = File(...),
    title: str = "Excel İle Yüklenen Yarışma",
    db: Session = Depends(get_db)
):
    """Parses an uploaded Excel file and creates a quiz."""
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(status_code=400, detail="Sadece .xlsx dosyaları kabul edilir.")

    user_cookie = request.cookies.get("user_session")
    if not user_cookie: raise HTTPException(status_code=401, detail="Not authenticated")
    
    current_user = db.query(models.User).filter(models.User.username == user_cookie).first()
    if not current_user: raise HTTPException(status_code=401, detail="User not found")

    try:
        contents = await file.read()
        wb = openpyxl.load_workbook(filename=BytesIO(contents), data_only=True)
        ws = wb.active
        
        # Validate headers
        header_row = [cell.value for cell in ws[1]]
        # Basic check if headers match roughly (at least first column)
        if not header_row or header_row[0] != TEMPLATE_HEADERS[0]:
            raise HTTPException(status_code=400, detail="Geçersiz şablon formatı. Lütfen sağlanan şablonu kullanın.")

        new_quiz = models.Quiz(
            title=title,
            description="Excel ile otomatik oluşturuldu",
            user_id=current_user.id,
            theme="standard",
            settings={"music_theme": "energetic"}
        )
        db.add(new_quiz)
        db.flush() # get ID

        # Iterate rows
        questions_created = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row[0]: # Skip empty question text
                continue
            
            # Unpack row
            # Expected: q_text, time, points, optA, optB, optC, optD, correct_letter
            q_text = str(row[0] or "").strip()
            time_limit = int(row[1] or 20)
            points = int(row[2] or 1000)
            
            # Options
            opt_a = str(row[3] or "").strip()
            opt_b = str(row[4] or "").strip()
            opt_c = str(row[5] or "").strip()
            opt_d = str(row[6] or "").strip()
            
            correct_char = str(row[7] or "").strip().upper()
            
            # Map correct char to index
            mapping = {'A': 0, 'B': 1, 'C': 2, 'D': 3}
            correct_idx = mapping.get(correct_char, -1) # -1 if invalid
            
            if correct_idx == -1 and (opt_a or opt_b):
                 raise HTTPException(status_code=400, detail=f"Satır {row[0]}: Geçersiz doğru cevap şıkkı '{correct_char}'. Lütfen sadece A, B, C veya D giriniz.")

            options_list = [
                {"text": opt_a, "is_correct": correct_idx == 0},
                {"text": opt_b, "is_correct": correct_idx == 1},
                {"text": opt_c, "is_correct": correct_idx == 2},
                {"text": opt_d, "is_correct": correct_idx == 3},
            ]
            
            question = models.Question(
                quiz_id=new_quiz.id,
                text=q_text,
                question_type="multiple_choice",
                time_limit=time_limit,
                points=points,
                options=options_list,
                image_url=None
            )
            db.add(question)
            questions_created += 1

        db.commit()
        db.refresh(new_quiz)
        
        return {"message": "Başarılı", "quiz_id": new_quiz.id, "questions_count": questions_created}

    except Exception as e:
        db.rollback()
        print(f"Import Error: {e}")
        raise HTTPException(status_code=500, detail=f"Dosya işlenirken hata: {str(e)}")
